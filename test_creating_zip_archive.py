import selene
import pytest
import csv
import os
from zipfile import ZipFile

from PyPDF2 import PdfReader
from openpyxl import load_workbook


@pytest.fixture()
def delete_test_zip():
    os.remove('resources/test.zip')


def test_creating_zip_archive(delete_test_zip):
    with ZipFile('resources/test.zip', mode='x') as newzip:
        newzip.write('resources/username.csv')
        newzip.write('resources/file_example_XLSX_50.xlsx')
        newzip.write('resources/docs-pytest-org-en-latest.pdf')
    check_csv()
    check_xlsx()
    check_pdf()


def check_csv():
    with open('resources/username.csv') as f:
        reader = csv.reader(f, delimiter=';')
        print('')
        for row in reader:
            pass
        assert reader.line_num == 7


def check_xlsx():
    workbook = load_workbook('resources/file_example_XLSX_50.xlsx')
    sheet = workbook.active
    assert f'{sheet.cell(row=2, column=2).value} {sheet.cell(row=2, column=3).value}' == 'Dulce Abril'
    workbook.close()


def check_pdf():
    reader = PdfReader('resources/docs-pytest-org-en-latest.pdf')
    page = reader.pages[0]
    text = page.extractText()
    assert 'Release 0.1' in text
